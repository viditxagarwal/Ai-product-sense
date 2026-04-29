"""Simulated execution engine for Phase 2 v1.

Reads the workflow graph, walks each node with realistic delays,
creates DB records (runs, steps, messages, files), and streams
events through a callback. Will be replaced by a real agent executor.

v2: Creates realistic Excel DCF models and markdown summaries.
    Handles subsequent runs with targeted edits based on user message keywords.
"""

import asyncio
import io
import logging
import random
import re
import traceback
import uuid
from datetime import datetime, timezone

import openpyxl

logger = logging.getLogger("ws.execution")
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import supabase

# ── Constants ─────────────────────────────────────────────────
COST_PER_TOKEN = 0.000003  # ~$3 per 1M tokens
STORAGE_BUCKET = "documents"

# ── DCF Model defaults ───────────────────────────────────────
DCF_DEFAULTS = {
    "wacc": 9.2,
    "terminal_growth": 2.5,
    "revenue_2025": 1_200_000,
    "revenue_growth": 12.0,
    "ebitda_margin": 22.0,
    "capex_pct": 5.0,
    "tax_rate": 25.0,
    "discount_rate": 9.2,
}

MODIFY_KEYWORDS = {"change", "update", "modify", "adjust", "revise", "set"}
APPEND_KEYWORDS = {"add", "append", "include", "insert", "extend"}


# ══════════════════════════════════════════════════════════════
# File creation helpers
# ══════════════════════════════════════════════════════════════

def _create_dcf_excel(params: dict | None = None) -> bytes:
    """Create a realistic DCF model Excel workbook and return as bytes."""
    p = {**DCF_DEFAULTS, **(params or {})}
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    currency_fmt = '#,##0'
    pct_fmt = '0.0%'
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    def _style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # ── Sheet 1: Assumptions ──────────────────────────────
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a.column_dimensions["A"].width = 25
    ws_a.column_dimensions["B"].width = 15

    assumptions = [
        ("Parameter", "Value"),
        ("WACC (%)", p["wacc"]),
        ("Terminal Growth Rate (%)", p["terminal_growth"]),
        ("Revenue 2025 ($)", p["revenue_2025"]),
        ("Revenue Growth (%)", p["revenue_growth"]),
        ("EBITDA Margin (%)", p["ebitda_margin"]),
        ("CapEx (% of Revenue)", p["capex_pct"]),
        ("Tax Rate (%)", p["tax_rate"]),
        ("Discount Rate (%)", p["discount_rate"]),
    ]
    for r, (label, val) in enumerate(assumptions, 1):
        ws_a.cell(row=r, column=1, value=label).border = thin_border
        ws_a.cell(row=r, column=2, value=val).border = thin_border
    _style_header(ws_a, 1, 2)

    # ── Sheet 2: Revenue Projections ──────────────────────
    ws_r = wb.create_sheet("Revenue")
    years = list(range(2025, 2030))
    ws_r.cell(row=1, column=1, value="Year")
    for ci, yr in enumerate(years, 2):
        ws_r.cell(row=1, column=ci, value=yr)
    _style_header(ws_r, 1, len(years) + 1)

    rev = p["revenue_2025"]
    revenues = []
    for yr in years:
        revenues.append(round(rev))
        rev *= 1 + p["revenue_growth"] / 100

    metrics = {
        "Revenue ($)": revenues,
        "EBITDA ($)": [round(r * p["ebitda_margin"] / 100) for r in revenues],
        "CapEx ($)": [round(r * p["capex_pct"] / 100) for r in revenues],
        "Tax ($)": [round(r * p["ebitda_margin"] / 100 * p["tax_rate"] / 100) for r in revenues],
        "Free Cash Flow ($)": [],
    }
    # FCF = EBITDA - CapEx - Tax
    metrics["Free Cash Flow ($)"] = [
        metrics["EBITDA ($)"][i] - metrics["CapEx ($)"][i] - metrics["Tax ($)"][i]
        for i in range(len(years))
    ]

    for ri, (label, vals) in enumerate(metrics.items(), 2):
        ws_r.cell(row=ri, column=1, value=label).border = thin_border
        for ci, v in enumerate(vals, 2):
            cell = ws_r.cell(row=ri, column=ci, value=v)
            cell.number_format = currency_fmt
            cell.border = thin_border
    ws_r.column_dimensions["A"].width = 22

    # ── Sheet 3: DCF Calculation ──────────────────────────
    ws_d = wb.create_sheet("DCF")
    ws_d.column_dimensions["A"].width = 28
    ws_d.column_dimensions["B"].width = 18

    fcf = metrics["Free Cash Flow ($)"]
    discount = p["discount_rate"] / 100

    pv_fcfs = [round(f / (1 + discount) ** (i + 1)) for i, f in enumerate(fcf)]
    terminal_value = round(fcf[-1] * (1 + p["terminal_growth"] / 100) / (discount - p["terminal_growth"] / 100))
    pv_terminal = round(terminal_value / (1 + discount) ** len(years))
    enterprise_value = sum(pv_fcfs) + pv_terminal

    dcf_rows = [
        ("Metric", "Value"),
        ("PV of FCF (Year 1)", pv_fcfs[0]),
        ("PV of FCF (Year 2)", pv_fcfs[1]),
        ("PV of FCF (Year 3)", pv_fcfs[2]),
        ("PV of FCF (Year 4)", pv_fcfs[3]),
        ("PV of FCF (Year 5)", pv_fcfs[4]),
        ("Terminal Value", terminal_value),
        ("PV of Terminal Value", pv_terminal),
        ("", ""),
        ("Enterprise Value", enterprise_value),
    ]
    for r, (label, val) in enumerate(dcf_rows, 1):
        ws_d.cell(row=r, column=1, value=label).border = thin_border
        cell = ws_d.cell(row=r, column=2, value=val)
        cell.border = thin_border
        if isinstance(val, (int, float)) and val:
            cell.number_format = currency_fmt
    _style_header(ws_d, 1, 2)
    # Highlight enterprise value
    ws_d.cell(row=10, column=2).font = Font(bold=True, size=12, color="1A7F37")

    # ── Sheet 4: Summary ──────────────────────────────────
    ws_s = wb.create_sheet("Summary")
    ws_s.column_dimensions["A"].width = 30
    ws_s.column_dimensions["B"].width = 20
    summary_rows = [
        ("Key Metric", "Value"),
        ("WACC", f"{p['wacc']}%"),
        ("Terminal Growth", f"{p['terminal_growth']}%"),
        ("5-Year Revenue CAGR", f"{p['revenue_growth']}%"),
        ("Average FCF Margin", f"{round(sum(fcf) / sum(revenues) * 100, 1)}%"),
        ("Enterprise Value ($)", f"${enterprise_value:,}"),
        ("Terminal Value (% of EV)", f"{round(pv_terminal / enterprise_value * 100, 1)}%"),
    ]
    for r, (label, val) in enumerate(summary_rows, 1):
        ws_s.cell(row=r, column=1, value=label).border = thin_border
        ws_s.cell(row=r, column=2, value=val).border = thin_border
    _style_header(ws_s, 1, 2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _create_analysis_md(params: dict | None = None) -> str:
    """Create a markdown analysis summary."""
    p = {**DCF_DEFAULTS, **(params or {})}
    return f"""# DCF Valuation Analysis Summary

## Overview
This analysis presents a Discounted Cash Flow (DCF) valuation model with a 5-year projection period (2025-2029).

## Key Assumptions
| Parameter | Value |
|-----------|-------|
| WACC | {p['wacc']}% |
| Terminal Growth Rate | {p['terminal_growth']}% |
| Base Revenue (2025) | ${p['revenue_2025']:,} |
| Revenue Growth | {p['revenue_growth']}% |
| EBITDA Margin | {p['ebitda_margin']}% |

## Methodology
1. **Revenue Projections**: Based on {p['revenue_growth']}% annual growth from a ${p['revenue_2025']:,} base
2. **EBITDA Calculation**: Applied {p['ebitda_margin']}% margin to projected revenues
3. **Free Cash Flow**: EBITDA less CapEx ({p['capex_pct']}% of revenue) and taxes ({p['tax_rate']}%)
4. **Terminal Value**: Gordon Growth Model with {p['terminal_growth']}% perpetual growth
5. **Discounting**: All cash flows discounted at WACC of {p['wacc']}%

## Key Findings
- The model projects steady revenue growth over the forecast period
- Free cash flow margins remain healthy throughout the projection
- Terminal value represents a significant portion of total enterprise value
- Sensitivity to WACC changes is moderate (±1% WACC = ~15% EV change)

## Risk Factors
- Revenue growth assumptions may be optimistic in a downturn
- EBITDA margins could compress with increased competition
- Terminal growth rate should not exceed long-term GDP growth

## Recommendation
Based on the analysis, the valuation appears reasonable under the current assumptions. We recommend performing sensitivity analysis on WACC and terminal growth rate before finalizing.

---
*Generated by AI Product Studio — DCF Analysis Module*
"""


def _upload_to_storage(thread_id: str, file_name: str, content: bytes, content_type: str) -> str:
    """Upload file to Supabase Storage and return the public URL."""
    path = f"threads/{thread_id}/{file_name}"
    try:
        supabase.storage.from_(STORAGE_BUCKET).upload(
            path, content, {"content-type": content_type, "upsert": "true"}
        )
        result = supabase.storage.from_(STORAGE_BUCKET).get_public_url(path)
        return result
    except Exception:
        # Fallback: return a placeholder URL if storage isn't configured
        return f"/storage/{STORAGE_BUCKET}/{path}"


def _detect_user_intent(message: str) -> str:
    """Detect whether the user wants to modify, append, or just query."""
    lower = message.lower()
    if any(kw in lower for kw in MODIFY_KEYWORDS):
        return "modify"
    if any(kw in lower for kw in APPEND_KEYWORDS):
        return "append"
    return "query"


def _extract_param_changes(message: str) -> dict[str, float]:
    """Extract parameter changes from user message. Returns dict of param_name -> new_value."""
    lower = message.lower()
    changes: dict[str, float] = {}

    # WACC
    m = re.search(r'wacc\s*(?:to|=|:)?\s*(\d+\.?\d*)\s*%?', lower)
    if m:
        changes["wacc"] = float(m.group(1))
        changes["discount_rate"] = float(m.group(1))  # WACC = discount rate

    # Terminal growth rate
    m = re.search(r'terminal\s+growth\s*(?:rate)?\s*(?:to|=|:)?\s*(\d+\.?\d*)\s*%?', lower)
    if m:
        changes["terminal_growth"] = float(m.group(1))

    # Revenue growth
    m = re.search(r'revenue\s+growth\s*(?:rate)?\s*(?:to|=|:)?\s*(\d+\.?\d*)\s*%?', lower)
    if m:
        changes["revenue_growth"] = float(m.group(1))

    # EBITDA margin
    m = re.search(r'ebitda\s+margin\s*(?:to|=|:)?\s*(\d+\.?\d*)\s*%?', lower)
    if m:
        changes["ebitda_margin"] = float(m.group(1))

    # Tax rate
    m = re.search(r'tax\s+rate\s*(?:to|=|:)?\s*(\d+\.?\d*)\s*%?', lower)
    if m:
        changes["tax_rate"] = float(m.group(1))

    # CapEx
    m = re.search(r'capex\s*(?:to|=|:)?\s*(\d+\.?\d*)\s*%?', lower)
    if m:
        changes["capex_pct"] = float(m.group(1))

    # Fallback: if "change" detected but no specific param, try generic percentage
    if not changes:
        m = re.search(r'(\d+\.?\d*)\s*%', message)
        if m:
            val = float(m.group(1))
            # Try to guess what parameter from context
            if "growth" in lower:
                if "terminal" in lower:
                    changes["terminal_growth"] = val
                else:
                    changes["revenue_growth"] = val
            elif "margin" in lower:
                changes["ebitda_margin"] = val
            elif "discount" in lower or "wacc" in lower:
                changes["wacc"] = val
                changes["discount_rate"] = val

    return changes


def _get_existing_files(thread_id: str) -> list[dict]:
    """Get existing AI-generated files for this thread."""
    result = supabase.table("thread_files").select("*").eq(
        "thread_id", thread_id
    ).eq("source", "ai_generated").execute()
    return result.data or []


def _get_latest_version(file_id: str) -> dict | None:
    """Get the latest version of a file."""
    result = supabase.table("file_versions").select("*").eq(
        "file_id", file_id
    ).order("version_number", desc=True).limit(1).execute()
    return result.data[0] if result.data else None


def _modify_excel_cells(thread_id: str, file_record: dict, step_id: str,
                        user_message: str, send_event_sync=None) -> dict:
    """Modify specific cells in an existing Excel file, create version + changes."""
    file_id = file_record["id"]
    latest_version = _get_latest_version(file_id)
    new_version_num = (latest_version["version_number"] + 1) if latest_version else 2

    # Determine what to change using the generic parameter extractor
    param_changes = _extract_param_changes(user_message)
    changes_to_make = []

    # Map of param_name -> (sheet!cell, row_in_assumptions, summary_row)
    PARAM_CELL_MAP = {
        "wacc": ("Assumptions!B2", "WACC (%)", "Summary!B2"),
        "terminal_growth": ("Assumptions!B3", "Terminal Growth Rate (%)", "Summary!B3"),
        "revenue_growth": ("Assumptions!B5", "Revenue Growth (%)", "Summary!B4"),
        "ebitda_margin": ("Assumptions!B6", "EBITDA Margin (%)", None),
        "capex_pct": ("Assumptions!B7", "CapEx (% of Revenue)", None),
        "tax_rate": ("Assumptions!B8", "Tax Rate (%)", None),
        "discount_rate": ("Assumptions!B9", "Discount Rate (%)", None),
    }

    for param, new_val in param_changes.items():
        cell_info = PARAM_CELL_MAP.get(param)
        if not cell_info:
            continue
        location, label, summary_cell = cell_info
        old_val = DCF_DEFAULTS.get(param, 0)

        changes_to_make.append({
            "location": location,
            "old_value": str(old_val),
            "new_value": str(new_val),
            "reason": f"User requested {label} adjustment to {new_val}%",
            "downstream_impact": {"affected_cells_count": 6, "affected_sheets": ["DCF", "Summary"]},
        })

        # Add downstream summary cell update if applicable
        if summary_cell:
            changes_to_make.append({
                "location": summary_cell,
                "old_value": f"{old_val}%",
                "new_value": f"{new_val}%",
                "reason": f"Summary updated to reflect new {label.split('(')[0].strip()}",
                "downstream_impact": None,
            })

    if not changes_to_make:
        # Random modifications if no specific value detected
        random_changes = [
            {
                "location": "Assumptions!B5",
                "old_value": str(DCF_DEFAULTS["revenue_growth"]),
                "new_value": str(round(random.uniform(8, 18), 1)),
                "reason": "Revenue growth adjusted based on updated market analysis",
                "downstream_impact": {"affected_cells_count": 10, "affected_sheets": ["Revenue", "DCF", "Summary"]},
            },
            {
                "location": "Assumptions!B6",
                "old_value": str(DCF_DEFAULTS["ebitda_margin"]),
                "new_value": str(round(random.uniform(18, 28), 1)),
                "reason": "EBITDA margin revised per latest operational data",
                "downstream_impact": {"affected_cells_count": 5, "affected_sheets": ["Revenue", "DCF"]},
            },
        ]
        changes_to_make = random.sample(random_changes, k=random.randint(1, 2))

    # Re-create the Excel file with new params
    params = dict(DCF_DEFAULTS)
    params.update(param_changes)

    excel_bytes = _create_dcf_excel(params)
    file_name = file_record["file_name"]
    file_url = _upload_to_storage(thread_id, file_name, excel_bytes,
                                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Update thread_files record
    supabase.table("thread_files").update({
        "file_url": file_url,
        "current_version": new_version_num,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }).eq("id", file_id).execute()

    # Create new file version
    version_record = supabase.table("file_versions").insert({
        "file_id": file_id,
        "version_number": new_version_num,
        "file_url": file_url,
        "operation_type": "targeted_edit",
        "change_summary": {"changes_count": len(changes_to_make), "primary_change": changes_to_make[0]["reason"]},
        "created_by": "ai",
        "trigger_step_id": step_id,
    }).execute()
    version_id = version_record.data[0]["id"]

    # Create file_changes records
    created_changes = []
    for ch in changes_to_make:
        change_record = supabase.table("file_changes").insert({
            "file_version_id": version_id,
            "change_type": "cell_modify",
            "location": ch["location"],
            "old_value": ch["old_value"],
            "new_value": ch["new_value"],
            "reason": ch["reason"],
            "downstream_impact": ch["downstream_impact"],
            "status": "pending",
        }).execute()
        created_changes.append(change_record.data[0])

    return {
        "file_id": file_id,
        "file_name": file_name,
        "version_number": new_version_num,
        "changes_count": len(changes_to_make),
        "changes": changes_to_make,
    }


def _append_to_file(thread_id: str, file_record: dict, step_id: str, user_message: str) -> dict:
    """Create an append version for a file."""
    file_id = file_record["id"]
    latest_version = _get_latest_version(file_id)
    new_version_num = (latest_version["version_number"] + 1) if latest_version else 2

    summary = {
        "action": "append",
        "description": f"Added new section based on: {user_message[:100]}",
        "content_added": "New analysis section with additional data points",
    }

    supabase.table("thread_files").update({
        "current_version": new_version_num,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }).eq("id", file_id).execute()

    supabase.table("file_versions").insert({
        "file_id": file_id,
        "version_number": new_version_num,
        "file_url": file_record["file_url"],
        "operation_type": "append",
        "change_summary": summary,
        "created_by": "ai",
        "trigger_step_id": step_id,
    }).execute()

    return {
        "file_id": file_id,
        "file_name": file_record["file_name"],
        "version_number": new_version_num,
        "operation": "append",
    }


# ══════════════════════════════════════════════════════════════
# Node simulators
# ══════════════════════════════════════════════════════════════

NODE_SIMULATORS: dict[str, callable] = {}


def _reg(node_type: str):
    def decorator(fn):
        NODE_SIMULATORS[node_type] = fn
        return fn
    return decorator


@_reg("agent_node")
def _sim_agent(node, **_):
    return {
        "output_payload": {"response": f"Analyzed input and produced structured output for '{node.get('data', {}).get('label', 'Node')}'."},
        "tokens": random.randint(800, 2000),
        "result_summary": "Generated analysis response",
    }


@_reg("route")
def _sim_route(node, **_):
    paths = ["path_a", "path_b", "default"]
    chosen = random.choice(paths)
    confidence = round(random.uniform(0.65, 0.98), 2)
    return {
        "output_payload": {"chosen_route": chosen},
        "routing_decision": {"candidates": paths, "chosen": chosen, "confidence": confidence},
        "confidence_score": confidence,
        "tokens": random.randint(100, 400),
        "result_summary": f"Routed to {chosen} (confidence: {confidence})",
    }


@_reg("retriever")
def _sim_retriever(node, **_):
    doc_count = random.randint(2, 5)
    docs = [{"title": f"Document {i+1}", "relevance": round(random.uniform(0.7, 0.99), 2)} for i in range(doc_count)]
    return {
        "output_payload": {"documents": docs, "total_retrieved": doc_count},
        "tokens": random.randint(200, 600),
        "result_summary": f"Retrieved {doc_count} documents from knowledge base",
    }


@_reg("calculator")
def _sim_calculator(node, **_):
    result = round(random.uniform(1000, 50000000), 2)
    return {
        "output_payload": {"calculation": "DCF Valuation", "result": result, "currency": "USD"},
        "tokens": random.randint(300, 800),
        "result_summary": f"Calculated value: ${result:,.2f}",
    }


@_reg("validator")
def _sim_validator(node, **_):
    checks = ["never_fabricate", "source_grounding", "calculation_accuracy"]
    results = {c: random.choice(["passed", "passed", "passed", "warning"]) for c in checks}
    return {
        "output_payload": {"guardrail_results": results},
        "guardrails_fired": [c for c, v in results.items() if v == "warning"],
        "tokens": random.randint(100, 300),
        "result_summary": f"Guardrails checked: {sum(1 for v in results.values() if v == 'passed')}/{len(checks)} passed",
    }


@_reg("file_writer")
def _sim_file_writer(node, **kwargs):
    """Enhanced: context-aware file operations."""
    context = kwargs.get("context", {})
    existing_files = context.get("existing_files", [])
    intent = context.get("intent", "query")

    if not existing_files:
        # First run: creation
        return {
            "output_payload": {"action": "file_creation", "files": ["DCF_Model.xlsx", "Analysis_Summary.md"]},
            "file_operation_type": "creation",
            "tokens": random.randint(500, 1500),
            "result_summary": "Created DCF_Model.xlsx and Analysis_Summary.md",
        }
    elif intent == "modify":
        return {
            "output_payload": {"action": "targeted_edit"},
            "file_operation_type": "targeted_edit",
            "tokens": random.randint(500, 1200),
            "result_summary": "Applied targeted modifications to existing files",
        }
    elif intent == "append":
        return {
            "output_payload": {"action": "append"},
            "file_operation_type": "append",
            "tokens": random.randint(300, 800),
            "result_summary": "Appended new content to existing files",
        }
    else:
        return {
            "output_payload": {"action": "no_file_change"},
            "file_operation_type": "none",
            "tokens": random.randint(200, 600),
            "result_summary": "Reviewed files without modification",
        }


@_reg("parallelization")
def _sim_parallel(node, **_):
    branch_count = node.get("data", {}).get("branchCount", 3)
    return {
        "output_payload": {"branches_executed": branch_count, "merge_method": "concatenate"},
        "tokens": random.randint(200, 500),
        "result_summary": f"Executed {branch_count} parallel branches",
    }


@_reg("loop")
def _sim_loop(node, **_):
    iterations = random.randint(1, 3)
    return {
        "output_payload": {"iterations": iterations, "exit_reason": "quality_threshold_met"},
        "tokens": random.randint(400, 1200),
        "result_summary": f"Looped {iterations} times, exited on quality threshold",
    }


@_reg("human_review")
def _sim_human(node, **_):
    return {
        "output_payload": {"status": "auto_approved", "note": "Simulated human approval"},
        "tokens": 0,
        "result_summary": "Human review auto-approved (simulated)",
    }


@_reg("end")
def _sim_end(node, **_):
    return {
        "output_payload": {"status": "workflow_complete"},
        "tokens": 0,
        "result_summary": "Workflow completed",
    }


# ── New component-model types ─────────────────────────────

@_reg("start")
def _sim_start(node, **_):
    return {
        "output_payload": {"status": "workflow_started"},
        "tokens": 0,
        "result_summary": "Workflow started",
    }


@_reg("node")
def _sim_node(node, **kwargs):
    """Unified node type — dispatches based on llmEnabled."""
    data = node.get("data", {})
    llm_enabled = data.get("llmEnabled", True)
    label = data.get("label", "Node")

    if llm_enabled:
        return {
            "output_payload": {"response": f"LLM analysis completed for '{label}'."},
            "tokens": random.randint(800, 2000),
            "result_summary": f"LLM node '{label}' generated response",
        }
    else:
        tools = data.get("boundTools", [])
        return {
            "output_payload": {"tools_executed": tools, "status": "tool_complete"},
            "tokens": random.randint(50, 200),
            "result_summary": f"Tool node '{label}' executed {len(tools)} tool(s)",
        }


@_reg("gate")
def _sim_gate(node, **_):
    """Gate node — human review checkpoint."""
    return {
        "output_payload": {"status": "auto_approved", "note": "Simulated gate approval"},
        "tokens": 0,
        "result_summary": "Gate auto-approved (simulated)",
    }


@_reg("split")
def _sim_split(node, **_):
    """Split node — parallel execution."""
    data = node.get("data", {})
    branch_count = data.get("branchCount", 3)
    merge_method = data.get("mergeMethod", "concatenate")
    return {
        "output_payload": {"branches_executed": branch_count, "merge_method": merge_method},
        "tokens": random.randint(200, 500),
        "result_summary": f"Split: {branch_count} branches, merged via {merge_method}",
    }


# ── Backward-compat aliases ──────────────────────────────
# Map old type names to the new simulators
NODE_SIMULATORS["step"] = _sim_node
NODE_SIMULATORS["classifier"] = _sim_node
NODE_SIMULATORS["human_checkpoint"] = _sim_gate
NODE_SIMULATORS["decision"] = _sim_route
NODE_SIMULATORS["parallel"] = _sim_split


def _default_sim(node, **_):
    return {
        "output_payload": {"result": "Step completed"},
        "tokens": random.randint(100, 500),
        "result_summary": "Step executed successfully",
    }


# ══════════════════════════════════════════════════════════════
# Graph ordering
# ══════════════════════════════════════════════════════════════

def _get_nodes_in_order(graph_data: dict) -> list[dict]:
    """Extract nodes from graph_data, attempt topological order via edges."""
    nodes = graph_data.get("nodes", [])
    edges = graph_data.get("edges", [])

    if not nodes:
        return []

    node_map = {n["id"]: n for n in nodes}
    in_degree = {n["id"]: 0 for n in nodes}
    adj = {n["id"]: [] for n in nodes}

    for e in edges:
        src = e.get("source")
        tgt = e.get("target")
        if src in adj and tgt in in_degree:
            adj[src].append(tgt)
            in_degree[tgt] = in_degree.get(tgt, 0) + 1

    # Kahn's algorithm
    queue = [nid for nid, deg in in_degree.items() if deg == 0]
    ordered = []
    while queue:
        nid = queue.pop(0)
        ordered.append(node_map[nid])
        for neighbor in adj.get(nid, []):
            in_degree[neighbor] -= 1
            if in_degree[neighbor] == 0:
                queue.append(neighbor)

    # Append any missed nodes (disconnected)
    seen = {n["id"] for n in ordered}
    for n in nodes:
        if n["id"] not in seen:
            ordered.append(n)

    return ordered


# ══════════════════════════════════════════════════════════════
# Main simulation entry point
# ══════════════════════════════════════════════════════════════

async def simulate_execution(thread_id: str, user_message: str, send_event):
    """Run the simulation. send_event is an async callable that sends a dict to the WebSocket."""
    logger.info("[exec] Starting simulation: thread=%s message='%s'", thread_id, user_message[:100])

    # Fetch thread + workflow
    try:
        thread = supabase.table("threads").select("*").eq("id", thread_id).single().execute()
    except Exception as e:
        logger.error("[exec] Failed to fetch thread %s: %s", thread_id, e)
        await send_event({"type": "error", "message": f"Failed to fetch thread: {e}"})
        return

    if not thread.data:
        logger.error("[exec] Thread not found: %s", thread_id)
        await send_event({"type": "error", "message": "Thread not found"})
        return

    logger.info("[exec] Thread found: workflow_id=%s", thread.data.get("workflow_id"))

    try:
        workflow = supabase.table("workflows").select("*").eq("id", thread.data["workflow_id"]).single().execute()
    except Exception as e:
        logger.error("[exec] Failed to fetch workflow: %s", e)
        await send_event({"type": "error", "message": f"Failed to fetch workflow: {e}"})
        return

    if not workflow.data:
        logger.error("[exec] Workflow not found: %s", thread.data["workflow_id"])
        await send_event({"type": "error", "message": "Workflow not found"})
        return

    graph_data = workflow.data.get("graph_data", {})
    nodes = _get_nodes_in_order(graph_data)
    logger.info("[exec] Graph has %d nodes, %d edges", len(graph_data.get("nodes", [])), len(graph_data.get("edges", [])))

    if not nodes:
        logger.info("[exec] No nodes in graph, using default agent node")
        nodes = [{"id": "default", "type": "agent_node", "data": {"label": "Default Agent"}}]

    # Check for existing files and determine user intent
    existing_files = _get_existing_files(thread_id)
    intent = _detect_user_intent(user_message)

    context = {
        "existing_files": existing_files,
        "intent": intent,
        "user_message": user_message,
        "thread_id": thread_id,
    }

    # Create execution run (user message already created by frontend via REST API)
    logger.info("[exec] Creating execution run for thread=%s", thread_id)
    try:
        run = supabase.table("execution_runs").insert({
            "thread_id": thread_id,
            "status": "running",
        }).execute()
    except Exception as e:
        logger.error("[exec] Failed to create execution run: %s\n%s", e, traceback.format_exc())
        await send_event({"type": "run_failed", "run_id": "", "error": f"Failed to create run: {e}"})
        return
    run_id = run.data[0]["id"]
    logger.info("[exec] Run created: run_id=%s", run_id)

    await send_event({
        "type": "run_started",
        "run_id": run_id,
        "step_count": len(nodes),
    })

    total_tokens = 0
    total_cost = 0.0
    total_duration = 0
    created_files = []
    modified_files = []

    try:
        for i, node in enumerate(nodes):
            node_type = node.get("type", "agent_node")
            node_name = node.get("data", {}).get("label", node_type)
            step_number = i + 1

            logger.info("[exec] Step %d/%d: type=%s name='%s'", step_number, len(nodes), node_type, node_name)

            # Create step record
            step = supabase.table("execution_steps").insert({
                "run_id": run_id,
                "step_number": step_number,
                "node_type": node_type,
                "node_name": node_name,
                "status": "running",
            }).execute()
            step_id = step.data[0]["id"]

            await send_event({
                "type": "step_started",
                "step_id": step_id,
                "step_number": step_number,
                "node_type": node_type,
                "node_name": node_name,
            })

            # Simulate thinking delay
            duration_ms = random.randint(500, 5000)
            chunks = random.randint(2, 4)
            per_chunk = duration_ms / chunks / 1000

            for c in range(chunks):
                await asyncio.sleep(per_chunk)
                progress_texts = [
                    f"Processing {node_name}...",
                    f"Analyzing input data...",
                    f"Generating output for step {step_number}...",
                    f"Applying {node_type} logic...",
                ]
                await send_event({
                    "type": "step_progress",
                    "step_id": step_id,
                    "content": progress_texts[c % len(progress_texts)],
                })

            # Generate simulated output (with context for file_writer)
            # Check componentType first (new model), fall back to node type (old model)
            component_type = node.get("data", {}).get("componentType", node_type)
            simulator = NODE_SIMULATORS.get(component_type) or NODE_SIMULATORS.get(node_type, _default_sim)
            sim_result = simulator(node, context=context)

            tokens = sim_result.get("tokens", 100)
            cost = round(tokens * COST_PER_TOKEN, 4)
            total_tokens += tokens
            total_cost += cost
            total_duration += duration_ms

            file_op = sim_result.get("file_operation_type", "none")

            # Update step as completed
            update_payload = {
                "status": "completed",
                "duration_ms": duration_ms,
                "tokens_used": tokens,
                "cost_usd": cost,
                "output_payload": sim_result.get("output_payload", {}),
                "routing_decision": sim_result.get("routing_decision", {}),
                "guardrails_fired": sim_result.get("guardrails_fired", []),
                "file_operation_type": file_op,
                "confidence_score": sim_result.get("confidence_score"),
            }
            supabase.table("execution_steps").update(update_payload).eq("id", step_id).execute()

            await send_event({
                "type": "step_completed",
                "step_id": step_id,
                "step_number": step_number,
                "duration_ms": duration_ms,
                "result_summary": sim_result.get("result_summary", "Done"),
                "file_operation_type": file_op,
            })

            # ── Handle file operations ────────────────────────
            if node_type == "file_writer" or file_op == "creation":
                if not existing_files:
                    # FIRST RUN: Create real Excel + Markdown files
                    created_files.extend(
                        await _create_initial_files(thread_id, step_id, send_event)
                    )
                elif file_op == "targeted_edit":
                    # SUBSEQUENT RUN: Modify existing Excel
                    excel_files = [f for f in existing_files if f["file_name"].endswith(".xlsx")]
                    for ef in excel_files:
                        mod_result = _modify_excel_cells(thread_id, ef, step_id, user_message)
                        modified_files.append(mod_result)
                        await send_event({
                            "type": "file_modified",
                            "file_id": mod_result["file_id"],
                            "file_name": mod_result["file_name"],
                            "file_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            "changes_count": mod_result["changes_count"],
                        })
                elif file_op == "append":
                    for ef in existing_files:
                        append_result = _append_to_file(thread_id, ef, step_id, user_message)
                        modified_files.append(append_result)
                        await send_event({
                            "type": "file_modified",
                            "file_id": append_result["file_id"],
                            "file_name": append_result["file_name"],
                            "operation": "append",
                        })

        # ── Complete the run ──────────────────────────────────
        supabase.table("execution_runs").update({
            "status": "completed",
            "total_duration_ms": total_duration,
            "total_tokens": total_tokens,
            "total_cost_usd": round(total_cost, 4),
            "step_count": len(nodes),
            "completed_at": datetime.now(timezone.utc).isoformat(),
        }).eq("id", run_id).execute()

        await send_event({
            "type": "run_completed",
            "run_id": run_id,
            "total_duration_ms": total_duration,
            "total_tokens": total_tokens,
            "total_cost_usd": round(total_cost, 4),
        })

        # ── Create assistant message ──────────────────────────
        assistant_content = _build_assistant_message(
            workflow.data.get("workflow_name", "workflow"),
            nodes, total_duration, total_tokens,
            created_files, modified_files, intent
        )

        supabase.table("thread_messages").insert({
            "thread_id": thread_id,
            "role": "assistant",
            "content": assistant_content,
            "message_type": "text",
            "metadata": {"run_id": run_id, "files": created_files},
        }).execute()

        await send_event({
            "type": "assistant_message",
            "content": assistant_content,
            "files": created_files,
        })

        logger.info("[exec] Run completed: run_id=%s duration=%dms tokens=%d", run_id, total_duration, total_tokens)

    except asyncio.CancelledError:
        logger.info("[exec] Run cancelled: run_id=%s", run_id)
        supabase.table("execution_runs").update({
            "status": "cancelled",
            "completed_at": datetime.now(timezone.utc).isoformat(),
        }).eq("id", run_id).execute()
        await send_event({"type": "run_failed", "run_id": run_id, "error": "Execution cancelled"})

    except Exception as e:
        logger.error("[exec] Run FAILED: run_id=%s error=%s\n%s", run_id, e, traceback.format_exc())
        supabase.table("execution_runs").update({
            "status": "failed",
            "completed_at": datetime.now(timezone.utc).isoformat(),
        }).eq("id", run_id).execute()
        await send_event({"type": "run_failed", "run_id": run_id, "error": str(e)})


async def _create_initial_files(thread_id: str, step_id: str, send_event) -> list[dict]:
    """Create Excel DCF model + Markdown summary on first run."""
    created = []

    # 1. Excel DCF Model
    excel_bytes = _create_dcf_excel()
    excel_name = "DCF_Model.xlsx"
    excel_url = _upload_to_storage(
        thread_id, excel_name, excel_bytes,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    excel_record = supabase.table("thread_files").insert({
        "thread_id": thread_id,
        "file_name": excel_name,
        "file_url": excel_url,
        "file_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "source": "ai_generated",
        "file_size_bytes": len(excel_bytes),
    }).execute()
    excel_id = excel_record.data[0]["id"]

    supabase.table("file_versions").insert({
        "file_id": excel_id,
        "version_number": 1,
        "file_url": excel_url,
        "operation_type": "creation",
        "created_by": "ai",
        "trigger_step_id": step_id,
    }).execute()

    created.append({"file_id": excel_id, "file_name": excel_name,
                     "file_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
    await send_event({
        "type": "file_created",
        "file_id": excel_id,
        "file_name": excel_name,
        "file_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    })

    # 2. Markdown Analysis Summary
    md_content = _create_analysis_md()
    md_name = "Analysis_Summary.md"
    md_bytes = md_content.encode("utf-8")
    md_url = _upload_to_storage(thread_id, md_name, md_bytes, "text/markdown")

    md_record = supabase.table("thread_files").insert({
        "thread_id": thread_id,
        "file_name": md_name,
        "file_url": md_url,
        "file_type": "text/markdown",
        "source": "ai_generated",
        "file_size_bytes": len(md_bytes),
    }).execute()
    md_id = md_record.data[0]["id"]

    supabase.table("file_versions").insert({
        "file_id": md_id,
        "version_number": 1,
        "file_url": md_url,
        "operation_type": "creation",
        "created_by": "ai",
        "trigger_step_id": step_id,
    }).execute()

    created.append({"file_id": md_id, "file_name": md_name, "file_type": "text/markdown"})
    await send_event({
        "type": "file_created",
        "file_id": md_id,
        "file_name": md_name,
        "file_type": "text/markdown",
    })

    return created


def _build_assistant_message(
    workflow_name: str, nodes: list, total_duration: int, total_tokens: int,
    created_files: list, modified_files: list, intent: str
) -> str:
    """Build a contextual assistant response message."""
    msg = (
        f"I've completed the analysis using the **{workflow_name}** workflow. "
        f"Processed {len(nodes)} steps in {total_duration / 1000:.1f}s using {total_tokens:,} tokens.\n\n"
    )

    if created_files:
        msg += "### Generated Files\n"
        for f in created_files:
            msg += f"- **{f['file_name']}** — "
            if f["file_name"].endswith(".xlsx"):
                msg += "DCF model with Assumptions, Revenue Projections, DCF Calculation, and Summary sheets\n"
            elif f["file_name"].endswith(".md"):
                msg += "Analysis narrative with methodology, key findings, and risk factors\n"
            else:
                msg += "Output file\n"

    if modified_files:
        msg += "\n### Modifications Made\n"
        for m in modified_files:
            if "changes_count" in m:
                msg += f"- **{m['file_name']}** (v{m['version_number']}) — {m['changes_count']} targeted change(s)\n"
                for ch in m.get("changes", []):
                    msg += f"  - `{ch['location']}`: {ch['old_value']} → {ch['new_value']} — *{ch['reason']}*\n"
            elif m.get("operation") == "append":
                msg += f"- **{m['file_name']}** (v{m['version_number']}) — Content appended\n"

    if not created_files and not modified_files and intent == "query":
        msg += "No file modifications were needed for this query. The existing files remain unchanged."

    return msg
